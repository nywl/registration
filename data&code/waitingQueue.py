#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Nov 21 22:09:37 2018

@author: pc
"""

#患者候诊队列变化

import xlrd
import time
from openpyxl import load_workbook
import dataProcess
import os
import xlsxwriter

realwaitingtime_yuyue = 1 #当该标识为1，则计算预约患者真实候诊时长，当该标识为0，则计算从服务开始时间到就诊的候诊时长
patients_records_regenerate = 1 #是否把患者记录从excel写入txt，写入的内容计算了患者的候诊时长
def datetime_get(s):
    if s.strip()!="":
        s_temp = s.split(" ")
        s_final = []
        s_final.append(s_temp[0])
        s_final.append(int(time.strftime("%H%M%S",time.strptime(s_temp[1], "%H:%M:%S"))))#日期类型数据长度相等，可直接字符串比对判断大小。时间类型长度不等（例如10：01：03比9：03：01字符串小），要转换成整数进行比较。
        return s_final
    else:
        return None
    
def time_revert(i):
    s = str(i)
    second = s[-2:]
    minute = s[-4:-2]
    hour = s[:-4]
    return hour+":"+minute+":"+second

#计算两个字符串时分秒之间的时间差    
def time_length(t1,t2):
    s1 = dataProcess.time_to_second(t1)
    s2 = dataProcess.time_to_second(t2)
    time_len = s2 - s1
#    s1 = time_revert(t1).split(":")
#    s2 = time_revert(t2).split(":")
#    time_len = (int(s2[0])-int(s1[0]))*3600+(int(s2[1])-int(s1[1]))*60+(int(s2[2])-int(s1[2]))
    return time_len

#候诊队列长度变化统计，仅考虑同一天的，不同日期的就诊应分开调用本函数
def houzhenLensStat(houzhen,t):    
    if t:#用参数t控制是否计算就诊时长，当现场和预约患者均考虑时，根据接诊时间判断每个人的就诊时长
        houzhen = jiuzhenTimeLength(houzhen)
    guahao_time_sorted = sorted(houzhen,key=lambda x:x[2])#使用匿名函数lambda,定义x，x[2]即list中每个list元素的第三个元素
    jiezhen_time_sorted = sorted(houzhen,key=lambda x:x[4])
    if t:#用参数t控制是否把排序后的结果写入文件，当预约、现场患者均考虑时才写入      
        f=open("./patients_guahao_time_sorted.txt","a+",encoding='utf-8')
        for i in range(0,len(guahao_time_sorted)):
            s = ""
            for j in range(0,len(guahao_time_sorted[i])):
                s = s + str(guahao_time_sorted[i][j])+","
            s = s[:-1]#去掉字符串末尾的逗号              
            f.write(s+"\n")
        f.flush
        f.close
        
        f=open("./patients_jiezhen_time_sorted.txt","a+",encoding='utf-8')
        for i in range(0,len(jiezhen_time_sorted)):
            s=""
            for j in range(0,len(jiezhen_time_sorted[i])):
                s = s + str(jiezhen_time_sorted[i][j])+","
            s = s[:-1]              
            f.write(s+"\n")
        f.flush
        f.close

    
    houzhen_count = len(houzhen)
    i = 0
    houzhen_lens = [] #存储候诊队列变化的list
    houzhen_len = 1
    for x in range(0,houzhen_count):
        if i+1>houzhen_count:
            break;
        else:
            while(jiezhen_time_sorted[x][4]>=guahao_time_sorted[i][2]):
                if i+1>=houzhen_count:
                    break;
                else:                    
                    houzhen_lens.append([guahao_time_sorted[i][1],guahao_time_sorted[i][2],houzhen_len,jiezhen_time_sorted[x][6],jiezhen_time_sorted[x][7]])
                    houzhen_len = houzhen_len+1
                i = i+1
                
        houzhen_len = houzhen_len-1
        houzhen_lens.append([jiezhen_time_sorted[x][3],jiezhen_time_sorted[x][4],houzhen_len,jiezhen_time_sorted[x][6],jiezhen_time_sorted[x][7]])
    return houzhen_lens

#计算就诊时长(该函数根据同一天就诊的患者的接诊时间，计算患者就诊时长，最后一个患者的就诊时长无法计算，用平均值替代或设为0)
def jiuzhenTimeLength(houzhen):
    jiuzhen_time_sorted = sorted(houzhen,key=lambda x:x[4])
    time_sum = 0
    for i in range(0,len(jiuzhen_time_sorted)-1):
        time_len = time_length(jiuzhen_time_sorted[i][4],jiuzhen_time_sorted[i+1][4])
        jiuzhen_time_sorted[i].insert(-2,time_len)#服务时间是上午还是下午的标识是新增的，放在原list末尾，防止对其他功能产生影响，所以就诊时长放在标识的前一个位置
        time_sum = time_sum + time_len
    #最后一名患者的就诊时长有多种计算方法，通过改变last_huanzhe的值，选择不同的方法
    last_huanzhe = 2
    #最后一个患者无法根据自己和下一个患者的接诊时间计算自己的就诊时长，故用该医生其他患者的平均就诊时长来估算
    if(last_huanzhe==1):    
        jiuzhen_time_sorted[len(jiuzhen_time_sorted)-1].insert(-2,int(time_sum/(len(jiuzhen_time_sorted)-1)))
    #最后一位患者就诊时长设为0，重排序时，不改变其就诊顺序，仍未最后一个就诊患者，此时可避免计算器真实就诊时长
    elif(last_huanzhe==2):
        jiuzhen_time_sorted[len(jiuzhen_time_sorted)-1].insert(-2,0)
    
    return jiuzhen_time_sorted

#把时分秒字符串转为秒
def time_to_second(t):
    second = int(t[-2:])
    minute = int(t[-4:-2])
    hour = int(t[:-4])
    s = hour * 3600 + minute * 60 + second
    return s

def main():
    data = xlrd.open_workbook('./outpatientRecord.xlsx')
    table = data.sheets()[0]
    nrows = table.nrows  #获取该sheet中的有效行数
    mat = []#把读取excel中内容放入list中
    
    for col in range(1, nrows): #第0行是字段名称，跳过
        mat.append(table.row_values(col))
    mat1 = sorted(mat, key=lambda x:(x[4], x[5], x[7], dataProcess.datetimeFormation(x[8])))
    houzhen = []  #定义list，存储一卡通号、挂号日期、挂号时间、接诊日期、接诊时间、号源名、专家名、类型、候诊时长、就诊时长
    houzhens = [] #list中每个元素为一个list，存储某位专家某个服务时段内对应的患者候诊队列变化情况
    houzhen_xianchang = [] #定义list，存储现场挂号患者一卡通号、挂号日期、挂号时间、接诊日期、接诊时间、号源名、专家名、类型、候诊时长、就诊时长
    houzhens_xianchang = []#list中每个元素为一个list，存储某位专家某个服务时段内对应的现场挂号患者候诊队列变化情况
    houzhen_yuyue = [] #定义list，存储预约患者一卡通号、挂号日期、挂号时间、接诊日期、接诊时间、号源名、专家名、类型、候诊时长、就诊时长
    houzhens_yuyue = []#list中每个元素为一个list，存储某位专家某个服务时段内对应的预约挂号患者候诊队列变化情况
    keshi= ""
    haoyuanming=""
    zhuanjia= ""
    fuwuqishishijian=0
    jiuzhen_date = ""
    #抽取需要的字段
    for col in range(0, len(mat1)): #原excel中第一行是字段名，故list中共有nrows-1行
        vars = mat1[col]
        if len(vars[3])==0 or len(vars[4])==0 or len(vars[5])==0 or len(vars[7])==0 or len(vars[8])==0:
            continue 
        guahao = dataProcess.datetime_get(vars[3])
        jiuzhen = dataProcess.datetime_get(vars[8])
        
        #时分秒格式化，由75440格式化为075440
        if len(guahao[1])==7:
            guahao[1] = "0" + guahao[1]
        if len(jiuzhen[1])==7:
            jiuzhen[1] = "0" + jiuzhen[1]
        guahao[0] = dataProcess.dateFormat(guahao[0])
        jiuzhen[0] = dataProcess.dateFormat(jiuzhen[0])
        
        fuwuqishishijian_new = 0
        if(len(vars[6].strip())!=0):
            fuwustart = int(vars[6].strip().split("-")[0].split(":")[0])
            if(fuwustart<13):
                fuwuqishishijian_new = 1
            else:
                fuwuqishishijian_new = 2
        #前面已把接诊时间row_temp[8]为空的给过滤掉，此处均不为空
        elif(len(jiuzhen)!=0):
            fuwustart = int(jiuzhen[1].split(":")[0])
            if(fuwustart<13):
                fuwuqishishijian_new = 1
            else:
                fuwuqishishijian_new = 2
        else:
            continue    
        
        if len(keshi)==0 and len(haoyuanming)==0 and len(jiuzhen)==0:
            keshi=vars[4]
            haoyuanming=vars[5]
            zhuanjia=vars[7]      
            jiuzhen_date=jiuzhen[0]
        #候诊时长的定义：现场患者是取号时间到接诊时间的间隔，预约患者是当接诊时间晚于预约时间时二者的间隔，当接诊时间早于预约时间设为0
        if len(guahao)!=0 and len(jiuzhen)!=0:
            if(vars[0].strip()=="appointment" and len(vars[6])>0 and realwaitingtime_yuyue==0):
                yuyue_starttime = vars[6].split("-")[0] + ":00"
                houzhenshichang = dataProcess.time_to_second(jiuzhen[1]) - dataProcess.time_to_second(yuyue_starttime)
                if(houzhenshichang<0):
                    houzhenshichang = 0            
            elif(vars[0].strip()=="on-site" or len(vars[6])==0 or realwaitingtime_yuyue==1):
                houzhenshichang = dataProcess.time_to_second(jiuzhen[1]) - dataProcess.time_to_second(guahao[1])
            
            if keshi!=vars[4] or haoyuanming!=vars[5] or zhuanjia!=vars[7] or fuwuqishishijian_new!=fuwuqishishijian or jiuzhen_date!=jiuzhen[0]:
                if houzhen:
                    houzhens.append(houzhen)               
                    houzhen = []
                if houzhen_xianchang:
                    houzhens_xianchang.append(houzhen_xianchang)
                    houzhen_xianchang = []
                if houzhen_yuyue:
                    houzhens_yuyue.append(houzhen_yuyue)
                    houzhen_yuyue = []
            houzhen.append([vars[1],guahao[0],guahao[1],jiuzhen[0],jiuzhen[1],vars[4],vars[7],vars[0],houzhenshichang,fuwuqishishijian_new,vars[5]])            
            if vars[0].strip() == "on-site":
                houzhen_xianchang.append([vars[1],guahao[0],guahao[1],jiuzhen[0],jiuzhen[1],vars[4],vars[7],vars[0],houzhenshichang,fuwuqishishijian_new,vars[5]])
                
            elif vars[0].strip() == "appointment":
                houzhen_yuyue.append([vars[1],guahao[0],guahao[1],jiuzhen[0],jiuzhen[1],vars[4],vars[7],vars[0],houzhenshichang,fuwuqishishijian_new,vars[5]])
            keshi = vars[4]
            haoyuanming=vars[5]
            zhuanjia = vars[7]
            fuwuqishishijian = fuwuqishishijian_new
            jiuzhen_date = jiuzhen[0]
    #一个houzhens中元素包含某类号源某专家的所有日期挂号接诊情况
    houzhens.append(houzhen)
    if houzhen_xianchang:
        houzhens_xianchang.append(houzhen_xianchang)
    if houzhen_yuyue:
        houzhens_yuyue.append(houzhen_yuyue)
    houzhen_lens = []
    for x in range(0,len(houzhens)): 
        houzhen_lens.append(houzhenLensStat(houzhens[x],patients_records_regenerate))
        

#openpyxl直接写如xlsx文件问报编码错误，加上下面一行注释即可通过
# coding: utf-8
    path_xlsx = "waitingQueue.xlsx"
    p1 = os.path.exists(path_xlsx)
    if(p1 is True):
        os.remove(path_xlsx)
    workbook = xlsxwriter.Workbook(path_xlsx)
    workbook.close()
    wb = load_workbook(path_xlsx)#openpyxl无法创建新的excel，只能读取已有的
    ws1 = wb.create_sheet(title="all") #新建sheet表并命令为“all”
    count = 0
    for e in houzhen_lens:       
        for i in range(0,len(e)):
        	for j in range(0,len(e[i])):
        		ws1.cell(count+1,j+1,str(e[i][j]))    #在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值
        	count = count + 1               
    #生成的excel中均为文本，通过excel自带value函数可把文本数据量转换为数值型，用于图表生成
    wb.save("waitingQueue.xlsx") 

#    vars = table.row_values(0)
#    for i in range(0,len(vars)-1):
#        for j in range(i+1,len(vars)):
#            s = "tab2 "+vars[i]+" "+vars[j]+",all"
#            f=file("D:/tab2-2013fd4.txt","a+")
#            f.writelines(s+"\n")
#    f.close


if __name__=="__main__":
    main()